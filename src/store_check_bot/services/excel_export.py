"""
Экспорт результатов отработки неучтённого товара в Excel.
"""

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from sqlalchemy.ext.asyncio import AsyncSession

from store_check_bot.repositories.products import get_verifications_for_export

STATUS_LABELS = {
    "processed": "Отработан",
    "not_processed": "Не отработан",
}

from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment


async def build_results_workbook(session: AsyncSession, export_dir: Path) -> Path:
    """Сформировать xlsx с отметками сотрудников."""
    export_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_path = export_dir / f"results_{timestamp}.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Результаты"

    # Заголовки
    headers = [
        "Отдел",
        "Артикул",
        "Наименование",
        "К-адрес",
        "Неучт. кол-во",
        "Дата вывода на проверку",
        "Статус",
        "Проверил (имя)",
        "Username",
        "Дата отметки",
    ]
    sheet.append(headers)

    # Данные
    rows = await get_verifications_for_export(session)
    for row in rows:
        dept, article, name, k_addr, unacc, shown_date, status, full_name, username, updated = row
        sheet.append(
            [
                dept,
                article,
                name,
                k_addr or "",
                unacc,
                shown_date.strftime("%d.%m.%Y") if shown_date else "",
                STATUS_LABELS.get(status, status),
                full_name,
                f"@{username}" if username else "",
                updated.strftime("%Y-%m-%d %H:%M") if updated else "",
            ]
        )

    # Автоматическое выравнивание столбцов по ширине
    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                if cell.value:
                    # Получаем длину значения в ячейке
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass

        # Устанавливаем ширину столбца с небольшим запасом
        adjusted_width = min(max_length + 2, 50)  # Максимум 50 символов
        sheet.column_dimensions[column_letter].width = adjusted_width

    # Опционально: центрирование для некоторых столбцов
    center_alignment = Alignment(horizontal='center', vertical='center')
    for col in ['A', 'B', 'E', 'J']:  # Отдел, Артикул, Неучт. кол-во, Дата отметки
        for row in range(2, sheet.max_row + 1):
            sheet[f'{col}{row}'].alignment = center_alignment

    workbook.save(file_path)
    return file_path
