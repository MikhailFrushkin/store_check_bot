"""
Импорт неучтённого товара из Excel (.xlsx) по формату «Неучтенные товары.xlsx».

Первая строка — заголовки на русском. При импорте каталог и история назначений сбрасываются.
"""

from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import delete
from sqlalchemy.ext.asyncio import AsyncSession

from store_check_bot.db.models import AppSettings, DailyAssignment, Product, Verification

# Заголовок Excel → поле Product
COLUMN_MAP: dict[str, str] = {
    "магазин": "shop",
    "отдел": "department",
    "подотдел": "subdepartment",
    "артикул": "article",
    "наименование": "name",
    "loc": "loc",
    "гамма": "gamma",
    "топ": "top",
    "теор. запас": "theoretical_stock",
    "вместимость": "capacity",
    "на ls": "on_ls",
    "на sscc в тз": "on_sscc",
    "кол-во на z адресе": "qty_z_address",
    "на rm": "on_rm",
    "на ем": "on_em",
    "на rd": "on_rd",
    "буфер склада": "warehouse_buffer",
    "неучтенный товар": "unaccounted_qty",
    "к-адрес": "k_address",
    "z-адрес": "z_address",
    "последнее движение": "last_movement",
}


def _normalize_header(value: object) -> str:
    """Привести заголовок к нижнему регистру."""
    if value is None:
        return ""
    return str(value).strip().lower()


def _to_int(value: object) -> int | None:
    """Преобразовать ячейку в int."""
    if value is None or str(value).strip() == "":
        return None
    return int(float(str(value).strip()))


def _to_float(value: object) -> float | None:
    """Преобразовать ячейку в float."""
    if value is None or str(value).strip() == "":
        return None
    return float(str(value).strip())


def _to_datetime(value: object) -> datetime | None:
    """Преобразовать ячейку в datetime."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%d.%m.%Y %H:%M", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def _cell_value(row: tuple, field: str, field_indices: dict[str, int]) -> object:
    """Прочитать значение ячейки по имени поля."""
    if field not in field_indices:
        return None
    return row[field_indices[field]]


def parse_products_from_excel(file_path: Path) -> list[dict[str, object]]:
    """
    Разобрать лист Excel в список словарей для Product.

    Raises:
        ValueError: Ошибка структуры или данных.
    """
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    sheet = workbook.active
    if sheet is None:
        raise ValueError("Файл не содержит листов")

    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) < 2:
        raise ValueError("Файл пуст или содержит только заголовок")

    headers = [_normalize_header(cell) for cell in rows[0]]
    field_indices: dict[str, int] = {}
    for idx, header in enumerate(headers):
        field = COLUMN_MAP.get(header)
        if field:
            field_indices[field] = idx

    required = {"department", "article", "name"}
    missing = required - set(field_indices)
    if missing:
        raise ValueError(
            f"Не найдены колонки: {', '.join(sorted(missing))}. "
            "Нужны: Отдел, Артикул, Наименование"
        )

    products: list[dict[str, object]] = []
    for row_num, row in enumerate(rows[1:], start=2):
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        try:
            department = _to_int(_cell_value(row, "department", field_indices))
        except (TypeError, ValueError) as exc:
            raise ValueError(f"Строка {row_num}: неверный отдел") from exc

        if department is None or department < 1 or department > 15:
            raise ValueError(f"Строка {row_num}: отдел должен быть от 1 до 15")

        article = str(_cell_value(row, "article", field_indices) or "").strip()
        name = str(_cell_value(row, "name", field_indices) or "").strip()
        if not article or not name:
            raise ValueError(f"Строка {row_num}: укажите артикул и наименование")

        item: dict[str, object] = {
            "shop": _to_int(_cell_value(row, "shop", field_indices)),
            "department": department,
            "subdepartment": _to_int(_cell_value(row, "subdepartment", field_indices)),
            "article": article,
            "name": name,
            "loc": str(_cell_value(row, "loc", field_indices) or "").strip() or None,
            "gamma": str(_cell_value(row, "gamma", field_indices) or "").strip() or None,
            "top": _to_int(_cell_value(row, "top", field_indices)),
            "theoretical_stock": _to_float(_cell_value(row, "theoretical_stock", field_indices)),
            "capacity": _to_float(_cell_value(row, "capacity", field_indices)),
            "on_ls": _to_float(_cell_value(row, "on_ls", field_indices)),
            "on_sscc": _to_float(_cell_value(row, "on_sscc", field_indices)),
            "qty_z_address": _to_float(_cell_value(row, "qty_z_address", field_indices)),
            "on_rm": _to_float(_cell_value(row, "on_rm", field_indices)),
            "on_em": _to_float(_cell_value(row, "on_em", field_indices)),
            "on_rd": _to_float(_cell_value(row, "on_rd", field_indices)),
            "warehouse_buffer": _to_float(_cell_value(row, "warehouse_buffer", field_indices)),
            "unaccounted_qty": _to_float(_cell_value(row, "unaccounted_qty", field_indices)),
            "k_address": str(_cell_value(row, "k_address", field_indices) or "").strip() or None,
            "z_address": str(_cell_value(row, "z_address", field_indices) or "").strip() or None,
            "last_movement": _to_datetime(_cell_value(row, "last_movement", field_indices)),
            "shown_for_check_date": None,
        }
        products.append(item)

    workbook.close()
    if not products:
        raise ValueError("В файле нет строк с товарами")
    return products


async def import_products_to_db(session: AsyncSession, products: list[dict]) -> int:
    """
    Полная замена каталога: товары, проверки, назначения на дни.

    Обновляет дату последней загрузки в app_settings.
    """
    await session.execute(delete(Verification))
    await session.execute(delete(DailyAssignment))
    await session.execute(delete(Product))

    app_settings = await session.get(AppSettings, 1)
    if app_settings is None:
        app_settings = AppSettings(id=1)
        session.add(app_settings)
    app_settings.last_excel_upload_at = datetime.utcnow()

    for item in products:
        session.add(Product(**item))

    await session.commit()
    return len(products)
